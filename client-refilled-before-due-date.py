import argparse
import sys
from pathlib import Path
import pandas as pd
import re
from datetime import timedelta
from typing import Dict, List, Tuple, Optional

# openpyxl imports for styling 
from openpyxl.styles import PatternFill

EXCEL_SHEET_ROW_LIMIT = 1_048_576  # Excel per-sheet limit

# RADET headers (exact columns expected inside the input files) 
RADET_COLS = {
    "state": "State",
    "lga": "L.G.A",
    "facility": "Facility Name",
    "patient_id": "Patient ID",
    "art_start": "ART Start Date (yyyy-mm-dd)",
    "art_status": "Current ART Status",
    "verification_outcome": "Client Verification Outcome",
}

# Pharmacy headers
DEFAULT_PHARM_ID_COL     = "Patient ID"
DEFAULT_PHARM_DATE_COL   = "Date Visit(yyyy-mm-dd)"
DEFAULT_PHARM_PERIOD_COL = "Refill Period"      # days
DEFAULT_PHARM_LINE_COL   = "Regimen Line"

# Allowed Regimen Lines
ALLOWED_REGIMEN_LINES = {
    "adult 1st line",
    "paediatric 1st line",
    "paediatric 2nd line",
    "arv prophylaxis for pregnant women",
    "adult 2nd line",
    "adult 3rd line",
    "arv prophylaxis for infants",
}

# ---------------------- CLI ----------------------
def parse_args():
    p = argparse.ArgumentParser(
        description="Flag multiple same-day refills and early refills across targeted IP paired structures."
    )
    p.add_argument(
        "--radet-dir", required=False,
        default=r"C:/Users/oluwabukola.arowolo/OneDrive - Palladium International, LLC/Documents/DataFi/DRM Slides/Refill-before-due-date/RADET",
        help="Path to folder containing your 7 RADET files."
    )
    p.add_argument(
        "--pharm-dir", required=False,
        default=r"C:/Users/oluwabukola.arowolo/OneDrive - Palladium International, LLC/Documents/DataFi/DRM Slides/Refill-before-due-date/Pharmacy report",
        help="Path to folder containing your 7 pharmacy files."
    )
    p.add_argument(
        "--out", required=False,
        default=r"C:/Users/oluwabukola.arowolo/OneDrive - Palladium International, LLC/Documents/DataFi/DRM Slides/Refill-before-due-date/output/refill_before_arv_elapsed3.xlsx",
        help="Path to output compiled Excel file (.xlsx)."
    )
    p.add_argument("--pharm-id-col",     default=DEFAULT_PHARM_ID_COL,   help="Pharmacy Patient Id column")
    p.add_argument("--pharm-date-col",   default=DEFAULT_PHARM_DATE_COL, help="Pharmacy visit date column")
    p.add_argument("--pharm-period-col", default=DEFAULT_PHARM_PERIOD_COL, help="Pharmacy refill period column (days)")
    p.add_argument("--pharm-line-col",   default=DEFAULT_PHARM_LINE_COL, help="Pharmacy regimen line column")
    p.add_argument("--dayfirst", action="store_true", help="Parse dates as D/M/Y.")
    p.add_argument("--visit-start", help="Optional visit start date (YYYY-MM-DD).")
    p.add_argument("--visit-end",   help="Optional visit end date (YYYY-MM-DD).")
    p.add_argument("--early-margin-days", type=int, default=30,
                   help="Flag an early refill only if next visit is at least this many days BEFORE due date.")
    p.add_argument("--recursive", action="store_true", help="Scan subfolders inside directories.")
    return p.parse_args()

# ---------------------- Helpers ----------------------
def _norm_cols(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    return out

def _read_table(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return pd.read_excel(path, dtype="object", engine="openpyxl")
    elif ext == ".xls":
        return pd.read_excel(path, dtype="object", engine="xlrd")
    elif ext == ".csv":
        return pd.read_csv(path, dtype="object")
    else:
        raise ValueError(f"Unsupported file type: {path}")

def _read_all_sheets(path: Path) -> Dict[str, pd.DataFrame]:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return pd.read_excel(path, dtype="object", engine="openpyxl", sheet_name=None)
    elif ext == ".xls":
        return pd.read_excel(path, dtype="object", engine="xlrd", sheet_name=None)
    else:
        raise ValueError(f"_read_all_sheets requires Excel file: {path}")

def _list_files(folder: Path, recursive: bool) -> List[Path]:
    patterns = ("*.xlsx", "*.xls", "*.csv")
    files: List[Path] = []
    if recursive:
        for pat in patterns:
            files.extend(folder.rglob(pat))
    else:
        for pat in patterns:
            files.extend(folder.glob(pat))
    return sorted(set(files))

def _to_dt(s: pd.Series, dayfirst: bool) -> pd.Series:
    return pd.to_datetime(s, errors="coerce", dayfirst=dayfirst)

def _parse_period_days(val) -> Optional[int]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        f = float(str(val).strip())
        d = int(round(f))
        return d if d > 0 else None
    except Exception:
        m = re.search(r"(\d+)", str(val))
        if m:
            d = int(m.group(1))
            return d if d > 0 else None
    return None

def _extract_ip_prefix(filename: str) -> str:
    """Extracts everything before the first underscore as the Implementing Partner name (e.g. 'ACE-1')."""
    if "_" in filename:
        return filename.split("_")[0].strip().upper()
    return Path(filename).stem.strip().upper()

# Processing Logic
def load_radet_file(f: Path, ip_name: str, dayfirst: bool) -> pd.DataFrame:
    """Loads a single RADET tracking sheet, applies 'Active' and 'Valid/Blank' filters."""
    df = _norm_cols(_read_table(f))
    
    missing = [c for c in RADET_COLS.values() if c not in df.columns]
    if missing:
        print(f"   ⚠️ Skipping file {f.name} because it is missing columns: {missing}", file=sys.stderr)
        return pd.DataFrame()
        
    # Filter 1: Current ART Status == Active & Active Restart
    status_col = RADET_COLS["art_status"]
    df = df[df[status_col].astype(str).str.strip().str.upper().isin(["ACTIVE", "ACTIVE RESTART"])]
    
    # Filter 2: Client Verification Outcome == Valid OR Blank/NaN
    verify_col = RADET_COLS["verification_outcome"]
    verify_clean = df[verify_col].astype(str).str.strip().str.upper()
    is_valid = verify_clean == "VALID"
    is_blank = (df[verify_col].isna()) | (df[verify_col].astype(str).str.strip() == "") | (verify_clean == "NAN")
    df = df[is_valid | is_blank]
    
    if df.empty:
        return pd.DataFrame()
        
    df["IP"] = ip_name
    df["_PID_"] = df[RADET_COLS["patient_id"]].astype(str).str.strip()
    df["_ART_START_"] = _to_dt(df[RADET_COLS["art_start"]], dayfirst)
    
    return df.drop_duplicates(subset=["_PID_"]).copy()

def process_pharmacy_file_into_map(
    f: Path, visits_map: Dict, radet_pid_set: set,
    id_col: str, date_col: str, period_col: str, line_col: str,
    dayfirst: bool, vstart: Optional[pd.Timestamp], vend: Optional[pd.Timestamp]
):
    """Processes a pharmacy file, parsing records *only* for PIDs active in this partner's RADET."""
    def _process_df(df: pd.DataFrame):
        df = _norm_cols(df)
        cols_l = {c.lower(): c for c in df.columns}
        lid, ldate, lper, lline = cols_l.get(id_col.lower()), cols_l.get(date_col.lower()), cols_l.get(period_col.lower()), cols_l.get(line_col.lower())
        if not (lid and ldate and lper and lline):
            return

        sub = df[[lid, ldate, lper, lline]].copy()
        sub["_PID_"] = sub[lid].astype(str).str.strip()
        
        sub = sub[sub["_PID_"].isin(radet_pid_set)]
        if sub.empty: return

        line_norm = sub[lline].astype(str).str.strip().str.casefold()
        sub = sub[line_norm.isin(ALLOWED_REGIMEN_LINES)]
        if sub.empty: return

        sub["_DATE_"]   = _to_dt(sub[ldate], dayfirst)
        sub["_PERIOD_"] = sub[lper].map(_parse_period_days)
        sub["_LINE_"]   = sub[lline].astype(str).str.strip()
        sub = sub[sub["_DATE_"].notna()]
        if sub.empty: return

        if vstart is not None or vend is not None:
            sub = sub[sub["_DATE_"].apply(lambda d: pd.isna(d) == False and (vstart is None or d >= vstart) and (vend is None or d <= vend))]
            if sub.empty: return

        for pid, d, prd, rline in zip(sub["_PID_"].values, sub["_DATE_"].values, sub["_PERIOD_"].values, sub["_LINE_"].values):
            lst = visits_map.get(pid, [])
            lst.append((pd.Timestamp(d), prd, rline))
            lst.sort(key=lambda x: x[0])
            if len(lst) > 3:
                lst[:] = lst[-3:]
            visits_map[pid] = lst

    if f.suffix.lower() in (".xlsx", ".xls"):
        for sname, sheet_df in _read_all_sheets(f).items():
            if sheet_df is not None and not sheet_df.empty:
                _process_df(sheet_df)
    elif f.suffix.lower() == ".csv":
        _process_df(_read_table(f))

def find_violations_for_pid_last_three(visits_last_three: List, early_margin_days: int) -> Tuple[List, List]:
    if not visits_last_three:
        return [], []
    visits_sorted = sorted(visits_last_three, key=lambda x: x[0])

    counts = {}
    for d, _, _ in visits_sorted:
        key = pd.Timestamp(d).normalize()
        counts[key] = counts.get(key, 0) + 1
    multiple_same_day = sorted([k for k, c in counts.items() if c >= 2])

    early_dates = []
    margin = timedelta(days=int(early_margin_days))
    for i in range(len(visits_sorted) - 1):
        d_i, p_i, _ = visits_sorted[i]
        d_next, _, _ = visits_sorted[i + 1]
        if p_i is None or pd.isna(p_i): continue
        due = d_i + timedelta(days=int(p_i))
        if d_next < (due - margin):
            early_dates.append(pd.Timestamp(d_next).normalize())

    return multiple_same_day, sorted(set(early_dates))

def assemble_outputs(radet_df: pd.DataFrame, visits_map: Dict, early_margin_days: int) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    def _fmt_date(d: pd.Timestamp) -> str:
        return pd.Timestamp(d).strftime("%Y-%m-%d")

    rows = []
    for pid, vlist in visits_map.items():
        if not vlist: continue
        v_sorted = sorted(vlist, key=lambda x: x[0])

        last3_dates_str    = "; ".join(_fmt_date(d) for d, _, _ in v_sorted)
        last3_periods_str  = "; ".join("" if (p is None or pd.isna(p)) else str(int(p)) for _, p, _ in v_sorted)
        last3_lines_str    = "; ".join("" if (ln is None or str(ln).strip() == "") else str(ln) for _, _, ln in v_sorted)
        latest_regimen_str = v_sorted[-1][2] if v_sorted[-1][2] is not None else ""

        # Calculate differences between refill dates
        if len(v_sorted) >= 2:
            refill_gaps = [
                str((v_sorted[i][0] - v_sorted[i - 1][0]).days)
                for i in range(1, len(v_sorted))
            ]
            refill_gaps_str = "; ".join(refill_gaps)
        else:
            refill_gaps_str = ""

        mult_dates, early_dates = find_violations_for_pid_last_three(v_sorted, early_margin_days)

        if mult_dates or early_dates:
            rows.append((
                pid,
                last3_dates_str,
                refill_gaps_str,
                last3_periods_str,
                last3_lines_str,
                latest_regimen_str,
                "; ".join(d.strftime("%Y-%m-%d") for d in mult_dates) if mult_dates else "",
                "; ".join(d.strftime("%Y-%m-%d") for d in early_dates) if early_dates else "",
                1 if early_dates else 0
            ))

    if not rows:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    viol_df = pd.DataFrame(rows, columns=[
    "_PID_",
    "Last 3 Refill Dates (asc)",
    "Refill Date Differences (days)",
    "Last 3 Refill Periods (days, asc)",
    "Last 3 Regimen Lines (asc)",
    "Regimen Line (latest among last 3)",
    "Multiple Refill Dates",
    "Early Refill Dates",
    "_Is_Early_Refill_"
    ])

    fields = [RADET_COLS["state"], RADET_COLS["lga"], RADET_COLS["facility"], RADET_COLS["patient_id"], RADET_COLS["art_start"]]
    
    out_detail = viol_df.merge(radet_df[["_PID_", "IP"] + fields], on="_PID_", how="left").rename(columns={
        RADET_COLS["state"]: "State", RADET_COLS["lga"]: "L.G.A",
        RADET_COLS["facility"]: "Facility Name", RADET_COLS["patient_id"]: "Patient Id",
        RADET_COLS["art_start"]: "ART Start Date (yyyy-mm-dd)"
    })

    out_agg = (out_detail.groupby(["State","L.G.A","Facility Name"], dropna=False)["Patient Id"]
               .nunique().reset_index().rename(columns={"Patient Id": "Count"})
               .sort_values(["State","L.G.A","Facility Name"]))

    out_summary = (out_detail[out_detail["_Is_Early_Refill_"] == 1]
                   .groupby(["IP", "State", "L.G.A", "Facility Name"], dropna=False)["Patient Id"]
                   .nunique().reset_index()
                   .rename(columns={"Patient Id": "Patients Flagged for Early Refills"})
                   .sort_values(["IP", "State", "L.G.A", "Facility Name"]))

    out_detail = out_detail.drop(columns=["_PID_", "_Is_Early_Refill_"])[
    ["IP","State","L.G.A","Facility Name","Patient Id",
     "ART Start Date (yyyy-mm-dd)",
     "Last 3 Refill Dates (asc)",
     "Refill Date Differences (days)",
     "Last 3 Refill Periods (days, asc)",
     "Last 3 Regimen Lines (asc)",
     "Regimen Line (latest among last 3)",
     "Multiple Refill Dates",
     "Early Refill Dates"]
    ]

    return out_detail, out_agg, out_summary

def apply_excel_highlights(file_path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(file_path)
    if "risk_summary" not in wb.sheetnames: return
    
    ws = wb["risk_summary"]
    count_col_idx = None
    for cell in ws[1]:
        if cell.value == "Patients Flagged for Early Refills":
            count_col_idx = cell.column
            break
            
    if count_col_idx is None: return

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row in range(2, ws.max_row + 1):
        val_cell = ws.cell(row=row, column=count_col_idx)
        try:
            val = int(val_cell.value) if val_cell.value is not None else 0
            if val > 100:
                val_cell.fill = red_fill
            elif 30 <= val <= 100:
                val_cell.fill = amber_fill
        except ValueError:
            continue

    wb.save(file_path)

# Execution flow 
def main():
    args = parse_args()
    radet_dir_path = Path(args.radet_dir)
    pharm_dir_path = Path(args.pharm_dir)
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if not radet_dir_path.exists(): raise NotADirectoryError(f"RADET folder path missing: {radet_dir_path}")
    if not pharm_dir_path.exists(): raise NotADirectoryError(f"Pharmacy folder path missing: {pharm_dir_path}")

    radet_files = _list_files(radet_dir_path, recursive=False)
    pharm_files = _list_files(pharm_dir_path, args.recursive)

    radet_by_ip: Dict[str, Path] = {_extract_ip_prefix(f.name): f for f in radet_files}
    pharm_by_ip: Dict[str, List[Path]] = {}
    for f in pharm_files:
        ip = _extract_ip_prefix(f.name)
        pharm_by_ip.setdefault(ip, []).append(f)

    print(f"Discovered {len(radet_by_ip)} distinct Implementing Partners inside your folders.")

    vstart = pd.to_datetime(args.visit_start) if args.visit_start else None
    vend   = pd.to_datetime(args.visit_end) if args.visit_end else None

    master_radet_list = []
    global_visits_map = {}

    for ip_name, radet_file in radet_by_ip.items():
        print(f"\nProcessing Cohort Group: {ip_name}")
        
        ip_radet_df = load_radet_file(radet_file, ip_name, args.dayfirst)
        if ip_radet_df.empty:
            print(f"  No Active/Valid data found in RADET for partner: {ip_name}")
            continue
            
        master_radet_list.append(ip_radet_df)
        ip_pid_set = set(ip_radet_df["_PID_"].tolist())
        print(f"   -> Loaded {len(ip_radet_df):,} target patients from RADET.")

        matching_pharm_files = pharm_by_ip.get(ip_name, [])
        if not matching_pharm_files:
            print(f"   Warning: No matching pharmacy file found for prefix '{ip_name}_'")
            continue

        for p_file in matching_pharm_files:
            print(f"   -> Analyzing pharmacy records from: {p_file.name}")
            process_pharmacy_file_into_map(
                p_file, global_visits_map, ip_pid_set,
                args.pharm_id_col, args.pharm_date_col, args.pharm_period_col, args.pharm_line_col,
                args.dayfirst, vstart, vend
            )

    if not master_radet_list:
        print("Error: No valid data frames remained after checking sheet structures and status filter rules.")
        return

    combined_radet_df = pd.concat(master_radet_list, ignore_index=True).drop_duplicates(subset=["_PID_"])

    print("\nStep 3: Calculating intervals and compiling multi-tier summary schemas...")
    out_detail, out_agg, out_summary = assemble_outputs(combined_radet_df, global_visits_map, early_margin_days=args.early_margin_days)

    if out_detail.empty:
        print("Complete. Zero tracking violations matched across parameter profiles.")
        return

    if len(out_detail) <= EXCEL_SHEET_ROW_LIMIT:
        with pd.ExcelWriter(out_path, engine="openpyxl", date_format="yyyy-mm-dd") as xw:
            out_summary.to_excel(xw, sheet_name="risk_summary", index=False)
            out_detail.to_excel(xw, sheet_name="violations_detail", index=False)
            out_agg.to_excel(xw, sheet_name="facility_aggregate", index=False)
        
        print("Step 4: Coloring cells based on threshold configurations...")
        apply_excel_highlights(out_path)
        print(f"Paired Analysis Workbook saved successfully: {out_path}")
    else:
        detail_csv = out_path.with_name(f"{out_path.stem}_violations_detail.csv")
        summary_csv = out_path.with_name(f"{out_path.stem}_risk_summary.csv")
        out_detail.to_csv(detail_csv, index=False)
        out_summary.to_csv(summary_csv, index=False)
        print(f"Output file exceeded Excel boundaries. Raw data split into CSV formats:\n - {detail_csv}\n - {summary_csv}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nERROR FAILURE: {e}", file=sys.stderr)
        sys.exit(1)